from openpyxl import *
import os
from dataclasses import dataclass


@dataclass
class Summary:
    white: float
    black: float
    dif: float
    white_trans: float


@dataclass
class Unit:
    price: float
    price_black: float
    price_white: float
    trans_price: float
    article: str
    count: int = 0


units = [Unit(0, 0, 0, 0, '')]
summary = Summary(0, 0, 0, 0)


for file in os.listdir('.'):
    #print(file)
    if str(file).find('.xlsx') != -1:
        wb = load_workbook(filename=file, data_only=True)
        sheet = wb['Sheet1']
        need_time = [4, 21]
        row = 2
        column = 1
        black_box = 'O'
        white_box = 'AF'
        black_profit = float(0)
        white_profit = float(0)
        while True:
            try:
                if int(sheet['L' + str(row)].value.split('-')[1]) > need_time[0] or \
                        int(sheet['L' + str(row)].value.split('-')[1]) == need_time[0] \
                        and int(sheet['L' + str(row)].value.split('-')[2]) >= need_time[1]:
                    summary.black += sheet[black_box+str(row)].value
                    summary.white += sheet[white_box+str(row)].value
                    summary.white_trans += sheet['AI' + str(row)].value
                    art_code = sheet['F'+str(row)].value
                    units_now = units
                    try:
                        temp_unit = [x for x in units_now if x.article == art_code]
                        art_index = units_now.index(temp_unit[0])
                        units[art_index].price += sheet[black_box + str(row)].value
                        units[art_index].price_black += sheet['P' + str(row)].value
                        units[art_index].price_white += sheet[white_box + str(row)].value
                        if sheet[black_box + str(row)].value == 0:
                            if sheet['K' + str(row)].value == 'Логистика':
                                units[art_index].trans_price += sheet['AI' + str(row)].value
                        else:
                            units[art_index].count += 1
                    except:
                        units.append(Unit(sheet[black_box + str(row)].value, sheet['P' + str(row)].value,
                                          sheet[white_box + str(row)].value, sheet['AI' + str(row)].value,
                                          sheet['F' + str(row)].value, 1))
                else:
                    pass
            except AttributeError:
                if sheet['K' + str(row)].value is not None:
                    pass
                else:
                    break
            row += 1


del(units[0])
wb = Workbook()
ws = wb.active
row = 1
column = 1
ws['A' + str(row)] = "Артикул товара"
ws['B' + str(row)] = "Цена розничная"
ws['C' + str(row)] = "Цена на WB"
ws['D' + str(row)] = "Кол-во штук"
ws['E' + str(row)] = "Цена минус комиссия"
ws['F' + str(row)] = "Стоимость логистики"
ws['G' + str(row)] = "Цена минус комиссия и логистика"
for unit in units:
    row += 1
    ws['A' + str(row)] = unit.article
    ws['B' + str(row)] = unit.price
    ws['C' + str(row)] = unit.price_black
    ws['D' + str(row)] = unit.count
    ws['E' + str(row)] = unit.price_white
    ws['F' + str(row)] = unit.trans_price
    ws['G' + str(row)] = unit.price_white-unit.trans_price

summary.dif = abs(((summary.white-summary.white_trans)/summary.black-1)*100)

ws['A' + str(row+1)] = "Итого сумма по рознице"
ws['D' + str(row+1)] = "Итого сумма минус комиссия"
ws['H' + str(row+1)] = "Итого сумма минус комиссия и логистика"
ws['A' + str(row+3)] = "Комиссия"
ws['A' + str(row+2)] = summary.black
ws['D' + str(row+2)] = summary.white
ws['H' + str(row+2)] = summary.white - summary.white_trans
ws['A' + str(row+4)] = summary.dif

wb.save('otchot.xlsx')
